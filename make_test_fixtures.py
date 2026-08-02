# -*- coding: utf-8 -*-
"""
테스트용 가짜 RA plan 생성기.

실물 파일은 Internal Use라 저장소에 넣지 않는다. 대신 실물과 **구조가 똑같은**
파일을 코드로 만들어 두어, 실물 없이도 엔진을 검증할 수 있게 한다.
구조 근거는 _reference/SOURCE_NOTES.md (실물에서 코드로 측정한 결과).

재현하는 구조:
  · 1행 안내 문구, 2행 형식 힌트, 3행 그룹 제목(병합), 4행 헤더, 5행부터 데이터
  · Responsible이 두 개 (A=구 담당자, B=현 담당자)
  · K~N = Pre-review, O~R = NDA or Variation
  · 날짜 칸에 N/A, TBD, '2027/01' 같은 실제 값들이 섞여 있음

    python make_test_fixtures.py
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT_DIR = Path(__file__).resolve().parent / "tests" / "fixtures"

# 실물 RA plan에서 판독한 행들. 마지막 몇 개는 경계 사례.
#  Responsible(구), Responsible(현), Product, Category, Project, Artwork, Status,
#  CP, dueA, dueB,
#  PreRev: PlanSub ActSub PlanApp ActApp | NDA: PlanSub ActSub PlanApp ActApp
ROWS = [
    ("RA01", "RA01", "Product-A", "Safety/Labelling",
     "RASP00001 Excipient and RASP00002 PSUSA update (Safety Category D)",
     None, "Completed", "Done", None, None,
     "2025-10-03", "2025-10-01", "2026-06-03", "2025-12-24",
     "2026-06-01", "2026-06-01", "2026-06-30", "2026-06-22"),

    ("RA01", "RA01", "Product-A", "CMC",
     "제형 기재사항 일부 삭제(local submission)",
     None, "Completed", "Done", None, None,
     "2025-10-03", "2025-10-01", "2026-06-03", "2025-11-10",
     "2026-06-01", "2026-06-01", "2026-06-30", "2026-06-22"),

    # ★ 검증 기준행 — 실물 간트: PRA Sub=2025-12-23, PRA App=2026-08-31
    #   App이 Actual(2026-05-06)이 아니라 Planned인 것이 핵심 (상태가 Ongoing).
    #   그리고 NDA Planned가 '2027/01'(일 없음)이라 partial로 잡혀야 한다.
    ("RA02", "RA02", "Product-B", "CMC Must-Win",
     "CR0000001 - New filling line for solvent at CMO site",
     None, "Ongoing", None, None, None,
     "2025-12-31", "2025-12-23", "2026-08-31", "2026-05-06",
     "2027/01", "TBD", "2027/01", "TBD"),

    ("RA03", "RA03", "Product-C", "Safety/Labelling",
     "Safety Labelling change Category B - Adverse event 1",
     None, "Completed", None, None, None,
     "2025-01-26", "TBD", "TBD", "TBD",
     "2025-01-26", "2025-01-24", "2026-01-26", "2026-01-26"),

    ("RA04", "RA04", "Product-D", "Safety/Labelling",
     "Safety Labelling change Category B - Adverse event 2",
     None, "Ongoing", None, None, None,
     "2025-09-30", "2025-08-29", "2026-03-31", "2025-11-20",
     "2026-04-30", "2026-04-10", "2026-12-10", "TBD"),

    # ★ 실물 간트: Sub=2026-09-01, App=2027-07-01 (Pending → Planned 사용)
    ("RA02", "RA02", "Product-E", "New indication",
     "STEP11 (BMI 25)", None, "Pending", None, None, None,
     "N/A", "N/A", "N/A", "N/A",
     "2026-09-01", "TBD", "2027-07-01", "TBD"),

    # ★ 실물 간트: Sub=2024-09-30, App=2025-05-06 (Completed → Actual 사용)
    ("RA02", "RA02", "Product-F", "NDA",
     "Orphan drug designation", None, "Completed", None, None, None,
     "N/A", "N/A", "N/A", "N/A",
     "2024-09-30", "2024-09-30", "2025-05-06", "2025-05-06"),

    # ★ 날짜가 전혀 없는 건 → 간트에서 제외되어야 한다
    ("RA02", "RA02", "Product-G", "Site addition",
     "Site addition for F/F", None, "Planned", None, None, None,
     "TBD", "TBD", "TBD", "TBD",
     "TBD", "TBD", "TBD", "TBD"),

    # ★ 담당자 칸에 미정 표기가 들어간 경우 → 빈칸으로 처리되어야 한다
    ("N/A", "N/A", "Product-H", "Renewal",
     "Renewal application 2027", None, "Planned", None, None, None,
     "N/A", "N/A", "N/A", "N/A",
     "2026 3Q", "TBD", "2027-06-30", "TBD"),
]

HEADERS = [
    "Responsible\n~20260607", "Responsible \n20260608~", "Product", "Category",
    "Project", "Artwork\nimpact", "Status", "CP delivery",
    "submission due date for safety update", "implementation date for CMC variation",
    "Planned Submission", "Actual Submission", "Planned Approval", "Actual Approval",
    "Planned Submission", "Actual Submission", "Planned Approval", "Actual Approval",
]

FORMAT_HINTS = {
    "H": "N/A, TBD, \nYYYY nQ, YYYY-MM-DD (15)",
    "K": "TBD, YYYY nQ, YYYY-MM-DD (15)",
    "M": "TBD, YYYY nQ, YYYY-MM-DD (15)",
    "O": "TBD, YYYY nQ, YYYY-MM-DD (15)",
    "Q": "TBD, YYYY nQ, YYYY-MM-DD (15)",
}


def _val(s):
    """'2025-10-03'은 진짜 날짜로, '2027/01'·'TBD'·'N/A'는 문자열 그대로."""
    if not isinstance(s, str) or not s:
        return s
    try:
        return date.fromisoformat(s)
    except ValueError:
        return s          # '2027/01', '2026 3Q', 'TBD', 'N/A' 등


def make_ra_plan(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RA plan"

    # 1행: 안내 문구 (실물처럼 병합)
    ws.merge_cells("H1:R1")
    ws["H1"] = "날짜는 반드시 YYYY-MM-DD로 작성; '최신의 plan 날짜'만 기재"

    # 2행: 형식 힌트
    for col, hint in FORMAT_HINTS.items():
        ws[f"{col}2"] = hint

    # 3행: 그룹 제목 (병합) — 이게 있어야 PRA/NDA가 갈린다
    ws.merge_cells("K3:N3")
    ws["K3"] = "Pre-review (사전검토 별도 하는 경우)"
    ws.merge_cells("O3:R3")
    ws["O3"] = "NDA or Variation\n(One step)"
    for c in ("K3", "O3"):
        ws[c].font = Font(bold=True)
        ws[c].alignment = Alignment(horizontal="center")

    # 4행: 헤더
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 5행부터 데이터
    for ri, row in enumerate(ROWS, start=5):
        for ci, v in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=_val(v))
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    from openpyxl.utils import get_column_letter

    widths = [13, 13, 14, 18, 50, 10, 12, 12, 16, 16, 13, 13, 13, 13, 13, 13, 13, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A4:R{4 + len(ROWS)}"
    ws.freeze_panes = "A5"
    wb.save(path)


if __name__ == "__main__":
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "sample_ra_plan.xlsx"
    make_ra_plan(out)
    print(f"생성됨: {out}")
    print("  (실물과 같은 구조: 헤더 4행, Responsible 2개, K~N=PreRev, O~R=NDA)")
