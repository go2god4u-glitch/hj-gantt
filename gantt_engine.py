# -*- coding: utf-8 -*-
"""
═══════════════════════════════════════════════════════════════════════════
 hj_gantt — RA plan(팀원 원장) → Gantt chart(매니저 뷰) 자동 변환 엔진
═══════════════════════════════════════════════════════════════════════════

무엇을 하는 물건인가
─────────────────────────────────────────────────────────────────────────
팀원들이 RA plan 엑셀에 일정을 적는다. 매니저는 간트로 본다. 지금까지는 그
사이를 사람 손이 이었다 — 날짜가 바뀌면 셀을 다시 칠하고, 새 프로젝트가
생기면 간트에도 손으로 옮겨 적고, 개정할 때마다 탭을 복제했다.

이 엔진은 그 손을 갈음한다. RA plan을 넣으면 **사용자의 기존 간트 양식**에
내용만 채워 간트를 통째로 다시 그린다. 그래서 —

    · 일정이 바뀌면        → 바가 따라 움직인다
    · 새 프로젝트가 생기면 → 간트에 그 행이 저절로 생긴다
    · RA plan에서 빠지면   → 간트에서도 사라진다

LLM을 쓰지 않는다. 전부 결정론적 코드이며 모든 판단 기준은 config.json에
노출돼 있다.

⚠️  간트가 이상하게 나왔다면 다음 순서로 보라
─────────────────────────────────────────────────────────────────────────
    [1] TROUBLESHOOTING.md   증상별 진단 — 대부분 config.json 한 줄로 끝난다
    [2] ARCHITECTURE.md      4단계 파이프라인 해부와 각 규칙의 실측 근거
    [3] _reference/SOURCE_NOTES.md
        실물 파일에서 **코드로 측정한** 구조 기록. 컬럼 배치, Category 분포,
        색상 규칙, 그리고 스크린샷만 보고 틀렸던 두 가지가 적혀 있다.

═══════════════════════════════════════════════════════════════════════════
[작전 일지] 사진으로 본 성(城)과 실제로 오른 성
                                                  (2026-08-02)
═══════════════════════════════════════════════════════════════════════════

1. 서막 — 사진 두 장으로 성을 그리다
   ─────────────────────────────────────────────────────────────────────
   처음 받은 것은 화면을 찍은 사진 두 장뿐이었다. 승상은 그것으로 성벽의
   높이와 문의 위치를 가늠하여 도면을 그렸다. 열여덟 행을 읽어내고, 제출일과
   승인일을 잇는 규칙을 세우고, 상태에 따라 색이 갈린다 하였다.

2. 첫 번째 균열 — 겹친 한 줄이 규칙을 심판하다
   ─────────────────────────────────────────────────────────────────────
   두 사진에 공교롭게 같은 행이 하나 겹쳐 있었다. 충전라인 신설 건.
   승상의 규칙은 2027-01-01을 내었으나 실제 간트는 2026-08-31이라 적혀 있었다.
   까닭을 파고드니 NDA란에 적힌 '2027/01'에 일(日)이 없었다. 연-월만 적힌
   것은 아직 정해지지 않았다는 표식이지 날짜가 아니었던 것.
   이를 물리고 다음 후보로 넘기니 비로소 맞아떨어졌다.

   → 한 줄이 여덟 개의 가정을 심판하였다. 승상은 이를 시험(test)에 새겨
     두었다. 규칙이 다시 어긋나면 그 시험이 먼저 울 것이다.

3. 그러나 — 성은 사진과 달랐다
   ─────────────────────────────────────────────────────────────────────
   이윽고 전하께서 실물 장부를 내리시니, 승상이 열어 보고 두 번 놀랐다.

   하나. Sub/App은 **한 쌍이 아니라 두 쌍**이었다.
         PRA(사전검토)와 NDA(본제출)가 나란히 놓여, RA plan의 두 그룹이
         간트의 두 쌍으로 그대로 건너갈 뿐이었다. 승상이 밤새 풀던
         "여덟 날짜 중 하나를 고르는 우선순위" — 그런 문제는 애초에
         존재하지 않았다. 사진의 오른편이 잘려 있었을 뿐이다.

   둘. 바의 색은 **상태가 아니라 Category**가 정하고 있었다.
       Others는 theme9, NDA는 theme7, Site addition은 theme6.
       진하기만이 준비기간과 검토기간을 갈랐다. 사진으로는 테마 색을
       분간할 수 없었으니, 승상은 보이지 않는 것을 지어내었던 것이다.

4. 교훈 — 색을 지어내지 아니한다
   ─────────────────────────────────────────────────────────────────────
   전하께서 이르시길, "색사도 모두 규칙이 있어 함부로 바꾸지 마라."
   하여 승상은 색을 만들어 내는 일을 그만두고, 대신 **전하의 양식에서 이미
   쓰이고 있는 색을 그대로 읽어다 쓰기로** 하였다(learn_palette).
   Category별로 어느 색이 쓰였는지, 연한 것과 진한 것이 무엇인지를 양식에서
   배워 그대로 돌려놓는다. 양식이 곧 법이요, 코드는 그 법을 따를 뿐이라.

5. 결론 — 모르는 것은 물어볼 곳을 코드에 심는다
   ─────────────────────────────────────────────────────────────────────
   아직 확정되지 않은 것이 남았다(준비기간의 길이 따위). 승상은 그것을
   아는 척하지 아니하고 `calibrate` 명을 두었다. 실물 RA plan과 실물 간트를
   나란히 놓으면 후보 규칙들을 스스로 대조하여 몇 %가 맞는지 답한다.

   추측을 코드에 새기지 말 것. 추측은 config.json에 두고, 실측할 길을
   함께 둘 것. 이것이 이번 작전의 요체(要諦)라.

═══════════════════════════════════════════════════════════════════════════
"""

from __future__ import annotations

import json
import re
import shutil
from copy import copy
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG_PATH = BASE_DIR / "config.json"

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_LOOKUP = {n.lower(): i + 1 for i, n in enumerate(MONTH_NAMES)}
MONTH_LOOKUP.update({f"{i}월": i for i in range(1, 13)})
MONTH_LOOKUP.update({str(i): i for i in range(1, 13)})

# 간트의 두 쌍. 이 순서가 곧 컬럼 순서다.
PHASES = ("pra", "nda")

# RA plan의 Category 드롭다운 목록 — 팀이 실제로 쓰는 17종, 적힌 순서 그대로.
#
# 간트는 이 값을 **그대로** 옮긴다. 접거나 바꾸거나 'Others'로 묶지 않는다.
# 예전에 20종을 7종으로 접었다가 팀원이 자기가 적은 값을 간트에서 찾지 못하는
# 일이 있었다. 그래서 원본 유지가 규칙이다.
#
# 이 목록의 쓰임은 검증 하나뿐이다: 데이터에 여기 없는 값(예: 'CMC Must-Win')이
# 있으면 '판독 참고'에 알려 준다. 손입력 오타이거나 목록이 낡았다는 신호다.
# 알려 주기만 하고 값은 절대 고치지 않는다.
RA_CATEGORIES = (
    "NDA", "New indication", "New posology", "Site addition", "CMC",
    "Safety/Labelling", "DMF", "New device", "Device variation", "Device GMP",
    "De-registration", "Launch", "Renewal", "Overseas site registration",
    "IND", "ODD", "Other",
)


def load_config(path: str | Path | None = None) -> dict:
    with open(Path(path) if path else DEFAULT_CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def _norm(v: Any) -> str:
    """
    셀 값을 사람이 읽는 문자열로.

    그냥 str()을 씌우면 안 된다. 엑셀에서 온 값은 문자열이 아닌 것이 섞여 있고,
    파이썬 기본 표현이 그대로 간트에 박힌다. 실제로 겪은 것들:

      · 셀 서식이 '시간'인 빈 칸  → time(0, 0)      → "00:00:00"
      · 날짜만 있는 칸           → datetime(자정)   → "2025-01-01 00:00:00"
      · 숫자로 저장된 제품 코드   → 1.0             → "1.0"

    Project 칸에 "00:00:00"이 찍혀 나온 것이 이 경우다. 값이 없다는 뜻이므로
    빈칸으로 돌린다. 시간이 실제로 들어 있을 때만 시:분을 남긴다.
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return (v.strftime("%Y-%m-%d") if v.time() == time(0, 0)
                else v.strftime("%Y-%m-%d %H:%M"))
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return "" if v == time(0, 0) else v.strftime("%H:%M")
    if isinstance(v, timedelta):
        return "" if v == timedelta(0) else str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))                     # 1.0 → "1"
    return re.sub(r"\s+", " ", str(v or "")).strip()


def _low(v: Any) -> str:
    return _norm(v).lower()


# ───────────────────────────────────────────────────────────────────────
# 날짜 판독
# ───────────────────────────────────────────────────────────────────────
_FULL = [
    re.compile(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$"),
    re.compile(r"^(\d{4})(\d{2})(\d{2})$"),
]
_YEAR_MONTH = re.compile(r"^(\d{4})[-/.](\d{1,2})$")
_YEAR_QUARTER = re.compile(r"^(\d{4})\s*([1-4])\s*[Qq]$")
_YEAR_MONTH_NAME = re.compile(r"^(\d{4})\s+([A-Za-z]{3,})$")
_YEAR_ONLY = re.compile(r"^(\d{4})$")


@dataclass
class DateValue:
    """
    셀 하나의 판독 결과.

    partial=True 는 '연-월/분기/연도만 있고 일(日)이 없다' = 아직 확정되지
    않았다는 뜻. 실물 RA plan 2행의 형식 안내가 `TBD, YYYY nQ, YYYY-MM-DD`
    이므로 이 표기들은 공식적으로 허용된 '미확정' 표기다.
    """
    value: date | None = None
    partial: bool = False
    raw: Any = None

    @property
    def usable(self) -> bool:
        return self.value is not None


def parse_cell_date(raw: Any, unresolved_tokens: Iterable[str]) -> DateValue:
    """엑셀 셀 하나를 날짜로 판독. TBD/N/A/줄바꿈/분기표기를 모두 견딘다."""
    if raw is None:
        return DateValue(raw=raw)
    if isinstance(raw, datetime):
        return DateValue(value=raw.date(), raw=raw)
    if isinstance(raw, date):
        return DateValue(value=raw, raw=raw)

    text = _norm(raw)
    if not text:
        return DateValue(raw=raw)

    tokens = {str(t).upper() for t in unresolved_tokens}
    if text.upper() in tokens:
        return DateValue(raw=raw)

    # '2026-12-31 (TBD)', '2022-9(E)', '2022-12-07(W)' → 괄호 주석 제거
    text = re.sub(r"\s*\([^)]*\)\s*$", "", text).strip()
    # '2024-11-30 2024-12-13' → 여러 날짜가 적힌 경우 첫 번째만
    if " " in text and re.match(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\s", text):
        text = text.split()[0]
    if not text or text.upper() in tokens:
        return DateValue(raw=raw)

    for pat in _FULL:
        m = pat.match(text)
        if m:
            try:
                return DateValue(value=date(*map(int, m.groups())), raw=raw)
            except ValueError:
                return DateValue(raw=raw)

    m = _YEAR_QUARTER.match(text)          # '2026 3Q'
    if m:
        y, q = int(m.group(1)), int(m.group(2))
        return DateValue(value=date(y, (q - 1) * 3 + 1, 1), partial=True, raw=raw)

    m = _YEAR_MONTH_NAME.match(text)       # '2025 Oct'
    if m:
        mo = MONTH_LOOKUP.get(m.group(2)[:3].lower())
        if mo:
            return DateValue(value=date(int(m.group(1)), mo, 1), partial=True, raw=raw)

    m = _YEAR_MONTH.match(text)            # '2026-12', '2027/01'
    if m:
        try:
            return DateValue(value=date(int(m.group(1)), int(m.group(2)), 1),
                             partial=True, raw=raw)
        except ValueError:
            return DateValue(raw=raw)

    m = _YEAR_ONLY.match(text)             # '2026'
    if m:
        return DateValue(value=date(int(m.group(1)), 1, 1), partial=True, raw=raw)

    return DateValue(raw=raw)


def add_months(d: date, months: int) -> date:
    """개월 가감. 말일 넘침은 그 달의 말일로 자른다 (3/31 -1개월 = 2/28)."""
    total = d.year * 12 + (d.month - 1) + months
    year, month = divmod(total, 12)
    month += 1
    nxt = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    last = (nxt - date(year, month, 1)).days
    return date(year, month, min(d.day, last))


def month_floor(d: date) -> date:
    return date(d.year, d.month, 1)


# ───────────────────────────────────────────────────────────────────────
# RA plan 판독
# ───────────────────────────────────────────────────────────────────────
@dataclass
class SourceColumns:
    header_row: int
    responsible: int | None = None
    responsible_alt: int | None = None
    product: int | None = None
    category: int | None = None
    project: int | None = None
    status: int | None = None
    pra_planned_sub: int | None = None
    pra_actual_sub: int | None = None
    pra_planned_app: int | None = None
    pra_actual_app: int | None = None
    nda_planned_sub: int | None = None
    nda_actual_sub: int | None = None
    nda_planned_app: int | None = None
    nda_actual_app: int | None = None
    notes: list[str] = field(default_factory=list)


def _hit(cell_text: Any, aliases: Iterable[str]) -> bool:
    """헤더에 개정일자·줄바꿈이 섞여 있어도 잡히도록 '포함'으로 본다."""
    t = _low(cell_text)
    return bool(t) and any(_low(a) and _low(a) in t for a in aliases)


def detect_source_columns(ws: Worksheet, config: dict, scan_rows: int = 25) -> SourceColumns:
    """
    RA plan 시트의 헤더 행과 컬럼 위치를 자동 탐지한다.

    날짜 컬럼은 이름만으로 PRA/NDA를 가릴 수 없다(양쪽 다 'Planned Submission').
    그래서 헤더 위 병합된 그룹 제목(`Pre-review …` / `NDA or Variation …`)의
    좌우 경계로 배정한다. 제목이 없으면 '앞 4개=PRA, 뒤 4개=NDA' 순서로 본다.
    """
    al = config["column_aliases"]
    max_col = min(ws.max_column or 1, 300)
    max_scan = min(ws.max_row or 1, scan_rows)
    plain = ("responsible", "product", "category", "project", "status")
    dated = ("planned_submission", "actual_submission",
             "planned_approval", "actual_approval")

    best_row, best = None, 0
    for r in range(1, max_scan + 1):
        score = sum(
            1 for c in range(1, max_col + 1)
            if any(_hit(ws.cell(row=r, column=c).value, al[k]) for k in plain + dated)
        )
        if score > best:
            best_row, best = r, score
    if best_row is None or best < 3:
        raise ValueError(
            "RA plan에서 헤더 행을 찾지 못했습니다. 상단 25행 안에 "
            "Responsible/Category/Project/Status 머리글이 있어야 합니다. "
            "이름이 다르면 config.json의 column_aliases에 추가하세요."
        )

    cols = SourceColumns(header_row=best_row)

    # Responsible이 여럿이면(구/현 담당자) config의 선호 표식이 붙은 쪽을 쓴다
    resp_hits = [c for c in range(1, max_col + 1)
                 if _hit(ws.cell(row=best_row, column=c).value, al["responsible"])]
    if resp_hits:
        prefer = config.get("responsible_prefer_marker", "20260608")
        chosen = next(
            (c for c in resp_hits
             if prefer and prefer in _norm(ws.cell(row=best_row, column=c).value)),
            resp_hits[-1],          # 표식이 없으면 가장 오른쪽 = 최신
        )
        cols.responsible = chosen
        others = [c for c in resp_hits if c != chosen]
        if others:
            cols.responsible_alt = others[0]
            cols.notes.append(
                f"Responsible 컬럼이 {len(resp_hits)}개 → "
                f"{get_column_letter(chosen)}열 사용 "
                f"(제외: {', '.join(get_column_letter(c) for c in others)})"
            )

    for key in ("product", "category", "project", "status"):
        for c in range(1, max_col + 1):
            if _hit(ws.cell(row=best_row, column=c).value, al[key]):
                setattr(cols, key, c)
                break

    # 그룹 경계
    spans: dict[str, tuple[int, int]] = {}
    for r in range(max(1, best_row - 3), best_row):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            for gkey, titles in config.get("group_titles", {}).items():
                if gkey in spans or not _hit(v, titles):
                    continue
                start = end = c
                for mr in ws.merged_cells.ranges:
                    if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                        start, end = mr.min_col, mr.max_col
                        break
                spans[gkey] = (start, end)

    short = {"planned_submission": "planned_sub", "actual_submission": "actual_sub",
             "planned_approval": "planned_app", "actual_approval": "actual_app"}
    hits = [(c, k) for c in range(1, max_col + 1)
            for k in dated if _hit(ws.cell(row=best_row, column=c).value, al[k])]

    if spans:
        for c, k in hits:
            for gkey, (s, e) in spans.items():
                if s <= c <= e:
                    attr = f"{gkey}_{short[k]}"
                    if getattr(cols, attr, "x") is None:
                        setattr(cols, attr, c)
                    break
    else:
        cols.notes.append("그룹 제목을 못 찾아 '앞 4개=PRA, 뒤 4개=NDA'로 배정했습니다.")
        for i, (c, k) in enumerate(hits):
            attr = f"{'pra' if i < 4 else 'nda'}_{short[k]}"
            if getattr(cols, attr, "x") is None:
                setattr(cols, attr, c)

    missing = [n for n in ("category", "status") if getattr(cols, n) is None]
    if missing:
        raise ValueError(
            f"RA plan 필수 컬럼을 못 찾았습니다: {', '.join(missing)} "
            f"(헤더 {best_row}행으로 판단). config.json의 column_aliases를 확인하세요."
        )
    return cols


# ───────────────────────────────────────────────────────────────────────
# 레코드
# ───────────────────────────────────────────────────────────────────────
@dataclass
class Record:
    source_row: int
    responsible: str
    product: str
    # RA plan에 적힌 Category. 간트에 그대로 찍히는 값은 언제나 이것 하나다.
    category: str
    # 색을 고르기 위해서만 쓰는 내부 키 (양식 팔레트가 접힌 이름으로 배워져
    # 있어서 필요하다). 셀에 쓰이지 않는다 — 이름이 'category'가 아닌 이유다.
    color_group: str
    project: str
    status: str
    pra_sub: date | None = None
    pra_app: date | None = None
    nda_sub: date | None = None
    nda_app: date | None = None
    pra_prep: date | None = None
    nda_prep: date | None = None
    # 승인일이 아직 없는 건의 '예상' 승인일. 바 길이에만 쓰고 날짜 칸에는 쓰지
    # 않는다 — 추정치를 확정 일정처럼 셀에 박아 넣으면 안 되기 때문이다.
    pra_app_est: date | None = None
    nda_app_est: date | None = None
    lead_months: int = 0
    origins: dict = field(default_factory=dict)
    flags: list[str] = field(default_factory=list)

    @property
    def has_dates(self) -> bool:
        return any((self.pra_sub, self.pra_app, self.nda_sub, self.nda_app))

    @property
    def first_sub(self) -> date | None:
        return next((d for d in (self.pra_sub, self.nda_sub) if d), None)

    def to_dict(self) -> dict:
        d = asdict(self)
        for k in ("pra_sub", "pra_app", "nda_sub", "nda_app", "pra_prep", "nda_prep",
                  "pra_app_est", "nda_app_est"):
            d[k] = d[k].isoformat() if d[k] else None
        return d

    def app_or_est(self, phase: str) -> date | None:
        """바를 어디까지 그릴지. 실제 승인일이 있으면 그것, 없으면 추정치."""
        return getattr(self, f"{phase}_app") or getattr(self, f"{phase}_app_est")

    @property
    def last_activity(self) -> date | None:
        """이 건의 마지막 시점. 추정 승인일까지 포함해서 본다."""
        ds = [d for d in (self.pra_sub, self.pra_app, self.nda_sub, self.nda_app,
                          self.pra_app_est, self.nda_app_est) if d]
        return max(ds) if ds else None


def _choose(actual: DateValue, planned: DateValue, reject_partial: bool,
            prefer_actual: bool = True) -> tuple[date | None, str]:
    """
    한 쌍(제출 또는 승인) 안에서 값을 고른다.

    prefer_actual=True  → Actual 우선 (이미 일어난 일이 사실이므로)
    prefer_actual=False → Planned 우선

    **제출일은 항상 Actual 우선**이지만 **승인일은 상태에 따라 갈린다.**
    아직 승인이 안 난 건(Ongoing/Planned/Pending)의 Actual Approval 칸에는
    중간 단계의 날짜가 적혀 있을 수 있어, 그 건의 '언제 승인될 예정인가'는
    Planned가 답이기 때문이다.

    > 검증: 기준행(충전라인 신설)(Ongoing) — PreRev Planned App=2026-08-31,
    >       Actual App=2026-05-06. 실제 간트의 PRA App은 2026-08-31(계획).
    >       상태가 Completed가 아니면 Planned를 먼저 봐야 재현된다.

    '2027/01'처럼 일(日)이 없는 값은 아직 확정되지 않았다는 표식이므로
    건너뛴다(reject_partial).
    """
    order = (("actual", actual), ("planned", planned))
    if not prefer_actual:
        order = (("planned", planned), ("actual", actual))

    skipped = []
    for label, dv in order:
        if not dv.usable:
            continue
        if reject_partial and dv.partial:
            skipped.append(f"{label}={_norm(dv.raw)}")
            continue
        return dv.value, label + (f"[건너뜀:{','.join(skipped)}]" if skipped else "")

    # 전부 미확정 표기뿐이면 그거라도 쓴다 (빈칸보다는 낫다)
    for label, dv in order:
        if dv.usable:
            return dv.value, f"{label}(미확정표기)"
    return None, "없음"


def _std_type(category: str, config: dict) -> str | None:
    """RA plan Category(20종) → 팀 표준 유형(5종). 기간 표를 찾는 열쇠."""
    low = _low(category)
    return {k.lower(): v for k, v in config.get("review_type_map", {}).items()}.get(low)


def prep_months(category: str, config: dict) -> int:
    """
    준비기간(연한 색 바) 길이. 자료 인도 → 제출까지 걸리는 개월 수.

    팀이 준 표 기준: NDA 4개월, 나머지 표준 유형은 3개월. 표에 없는 Category는
    default_prep_months.
    """
    std = _std_type(category, config)
    table = {k.lower(): v for k, v in config.get("prep_months_by_type", {}).items()}
    if std and _low(std) in table:
        return int(table[_low(std)])
    return int(config.get("default_prep_months", 3))


def review_months(category: str, config: dict) -> int:
    """
    승인일이 아직 없는 건의 검토기간(개월). 팀이 준 표준 소요기간 표를 쓴다.

    RA plan의 Category는 20종인데 표는 5개 표준 유형으로 되어 있다. 그래서
    review_type_map으로 한 번 접은 뒤 개월 수를 찾는다. 표에 없는 유형은
    default_review_months.
    """
    std = _std_type(category, config)
    table = {k.lower(): v for k, v in config.get("review_months_by_type", {}).items()}
    if std and _low(std) in table:
        return int(table[_low(std)])
    return int(config.get("default_review_months", 8))


def build_records(ws: Worksheet, cols: SourceColumns, config: dict) -> list[Record]:
    tokens = config.get("unresolved_tokens", ["TBD"])
    reject = bool(config.get("reject_partial_dates", True))
    # ⚠️ 아래 셋은 **색 그룹**을 정하는 규칙이지 Category 라벨을 바꾸는 규칙이
    # 아니다. 간트에 찍히는 Category는 언제나 RA plan 원본 그대로다.
    passthrough = {c.lower() for c in config.get("palette_group_passthrough", [])}
    explicit = {k.lower(): v for k, v in config.get("palette_group_map", {}).items()}
    fallback = config.get("palette_group_fallback", "Others")
    name_mode = config.get("project_name_mode", "verbatim")
    max_chars = int(config.get("project_name_max_chars", 60))

    out: list[Record] = []
    for r in range(cols.header_row + 1, (ws.max_row or cols.header_row) + 1):
        def cv(attr):
            c = getattr(cols, attr)
            return ws.cell(row=r, column=c).value if c else None

        project = _norm(cv("project"))
        cat_src = _norm(cv("category"))
        status = _norm(cv("status"))
        product = _norm(cv("product"))
        if not (project or cat_src or status or product):
            continue

        # 승인일은 '완료된 건'만 Actual을 먼저 본다 (_choose docstring 참조)
        completed = status.lower() in {
            s.lower() for s in config.get("completed_statuses", ["Completed"])}

        origins: dict[str, str] = {}
        dates: dict[str, date | None] = {}
        for ph in PHASES:
            for kind in ("sub", "app"):
                a = parse_cell_date(cv(f"{ph}_actual_{kind}"), tokens)
                p = parse_cell_date(cv(f"{ph}_planned_{kind}"), tokens)
                prefer_actual = True if kind == "sub" else completed
                val, origin = _choose(a, p, reject, prefer_actual)
                dates[f"{ph}_{kind}"] = val
                origins[f"{ph}_{kind}"] = origin

        # 준비기간(연한 색): 자료 인도 → 제출. 표준 유형별 개월 수만큼 제출일
        # 앞으로 물린다. PRA·NDA 두 구간에 같은 길이를 적용한다.
        lead = prep_months(cat_src, config)
        prep = {
            f"{ph}_prep": (add_months(dates[f"{ph}_sub"], -lead)
                           if dates[f"{ph}_sub"] else None)
            for ph in PHASES
        }

        # 승인일이 비어 있으면 표준 검토기간으로 끝점을 추정한다. 이게 없으면
        # 예정 건들이 제출월 한 칸짜리 점으로만 찍혀 사실상 안 보인다.
        # PRA는 유형과 무관하게 팀이 정한 고정값(기본 5개월)을 쓴다.
        pra_rev = int(config.get("pra_review_months", 5))
        nda_rev = review_months(cat_src, config)
        est = {}
        for ph in PHASES:
            sub, app_ = dates[f"{ph}_sub"], dates[f"{ph}_app"]
            months = pra_rev if ph == "pra" else nda_rev
            est[f"{ph}_app_est"] = (add_months(sub, months)
                                    if sub and not app_ else None)

        # 색 그룹만 정한다. cat_src(원본 Category)는 여기서 절대 바뀌지 않는다.
        low = cat_src.lower()
        if low in explicit:
            color_group = explicit[low]
        elif low in passthrough:
            color_group = cat_src
        else:
            color_group = fallback

        display = project
        if name_mode == "truncate" and len(project) > max_chars:
            display = project[: max_chars - 1].rstrip() + "…"

        flags = []
        for ph in PHASES:
            s, a = dates[f"{ph}_sub"], dates[f"{ph}_app"]
            if s and a and a < s:
                flags.append(f"{ph.upper()} 승인일({a})이 제출일({s})보다 빠름")

        # 'N/A', 'done' 같은 미확정 표기가 담당자 칸에 적혀 있기도 하다 — 빈칸으로 본다
        token_set = {str(t).upper() for t in tokens}
        resp = _norm(cv("responsible"))
        if resp.upper() in token_set:
            resp = ""

        out.append(Record(
            source_row=r,
            responsible=resp,
            product=product,
            category=cat_src,
            color_group=color_group,
            project=display,
            status=status,
            lead_months=lead,
            origins=origins,
            flags=flags,
            **dates, **prep, **est,
        ))
    return out


def read_ra_plan(path: str | Path, config: dict,
                 sheet_name: str | None = None) -> tuple[list[Record], SourceColumns, str]:
    wb = load_workbook(path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws, best = None, -1
        for name in wb.sheetnames:
            cand = wb[name]
            try:
                detect_source_columns(cand, config)
            except ValueError:
                continue
            if (cand.max_row or 0) > best:
                ws, best = cand, cand.max_row or 0
        if ws is None:
            raise ValueError(
                "어느 시트에서도 RA plan 헤더를 찾지 못했습니다. "
                "시트명을 지정하거나 config.json의 column_aliases를 확인하세요."
            )
    cols = detect_source_columns(ws, config)
    records = build_records(ws, cols, config)

    # RA plan의 Category 드롭다운 목록에 없는 값이 섞여 있으면 알려 준다.
    # 값을 고치지는 않는다 — 간트에는 적힌 그대로 나가야 하기 때문이다.
    # 다만 'CMC Must-Win'처럼 목록에 없는 표기는 손입력 오타이거나 목록이
    # 낡았다는 신호라서, 사람이 한 번 보고 판단할 수 있게 남긴다.
    known = {_low(c) for c in config.get("known_categories") or RA_CATEGORIES}
    if known:
        unknown = sorted({r.category for r in records
                          if r.category and _low(r.category) not in known})
        if unknown:
            cols.notes.append(
                "Category 목록에 없는 값 " + str(len(unknown)) + "종: "
                + ", ".join(unknown) + " (값은 그대로 간트에 씁니다)")
    return records, cols, ws.title


# ───────────────────────────────────────────────────────────────────────
# 간트 양식(템플릿) 해석
# ───────────────────────────────────────────────────────────────────────
@dataclass
class TemplateLayout:
    sheet_name: str
    header_row: int
    first_data_row: int
    last_data_row: int
    month_cols: dict[int, date]
    info_cols: dict[str, int]
    notes: list[str] = field(default_factory=list)


def detect_template_layout(ws: Worksheet, config: dict, scan_rows: int = 30) -> TemplateLayout:
    """
    사용자 간트 양식에서 '어디에 무엇을 쓸지'를 역탐지한다.

    Sub/App은 두 쌍(PRA / NDA or variations)이므로, 헤더 위의 그룹 제목으로
    어느 Sub이 어느 쌍인지 가른다. 이걸 놓치면 사전검토 일정이 본제출 칸에
    들어간다 — 실물 파일 판독에서 실제로 저질렀던 실수다.
    """
    tal = config.get("template_aliases", {})
    max_col = min(ws.max_column or 1, 400)
    max_scan = min(ws.max_row or 1, scan_rows)

    month_row, best = None, 0
    for r in range(1, max_scan + 1):
        hits = sum(1 for c in range(1, max_col + 1)
                   if _low(ws.cell(row=r, column=c).value) in MONTH_LOOKUP)
        if hits > best:
            month_row, best = r, hits
    if month_row is None or best < 6:
        raise ValueError("양식에서 월(Jan~Dec) 헤더 행을 찾지 못했습니다.")

    # 연도 행
    year_row, best_y = None, 0
    for r in range(max(1, month_row - 4), month_row):
        cnt = sum(1 for c in range(1, max_col + 1)
                  if re.fullmatch(r"(19|20)\d{2}", _norm(ws.cell(row=r, column=c).value)))
        if cnt > best_y:
            year_row, best_y = r, cnt

    year_by_col: dict[int, int] = {}
    if year_row:
        cur = None
        for c in range(1, max_col + 1):
            t = _norm(ws.cell(row=year_row, column=c).value)
            if re.fullmatch(r"(19|20)\d{2}", t):
                cur = int(t)
            if cur:
                year_by_col[c] = cur
        for mr in ws.merged_cells.ranges:      # 병합 구간 전체에 같은 연도
            if mr.min_row <= year_row <= mr.max_row:
                t = _norm(ws.cell(row=mr.min_row, column=mr.min_col).value)
                if re.fullmatch(r"(19|20)\d{2}", t):
                    for c in range(mr.min_col, mr.max_col + 1):
                        year_by_col[c] = int(t)

    notes: list[str] = []
    month_cols: dict[int, date] = {}
    rolling, prev = None, None
    for c in range(1, max_col + 1):
        m = MONTH_LOOKUP.get(_low(ws.cell(row=month_row, column=c).value))
        if not m:
            continue
        y = year_by_col.get(c)
        if y is None:
            if rolling is None:
                notes.append("연도 행을 못 찾아 첫 월을 올해로 가정했습니다.")
                rolling = date.today().year
            elif prev is not None and m < prev:
                rolling += 1
            y = rolling
        rolling, prev = y, m
        month_cols[c] = date(y, m, 1)
    if not month_cols:
        raise ValueError("양식의 월 컬럼을 해석하지 못했습니다.")

    first_month_col = min(month_cols)

    # 그룹 제목(PRA / NDA or variations)의 좌우 경계
    spans: dict[str, tuple[int, int]] = {}
    for r in range(max(1, month_row - 3), month_row):
        for c in range(1, first_month_col):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            for gkey in PHASES:
                titles = tal.get(f"{gkey}_group", [])
                if gkey in spans or not titles or not _hit(v, titles):
                    continue
                s = e = c
                for mr in ws.merged_cells.ranges:
                    if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                        s, e = mr.min_col, mr.max_col
                        break
                spans[gkey] = (s, e)

    info: dict[str, int] = {}
    header_row = month_row
    for r in range(max(1, month_row - 2), month_row + 1):
        found: dict[str, int] = {}
        for c in range(1, first_month_col):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            for key in ("responsible", "category", "product", "project", "status"):
                if key not in found and _hit(v, tal.get(key, [])):
                    found[key] = c
            for kind in ("sub", "app"):
                if not _hit(v, tal.get(kind, [])):
                    continue
                ph = next((g for g, (s, e) in spans.items() if s <= c <= e), None)
                if ph is None:  # 그룹 제목이 없으면 왼쪽부터 pra, nda 순
                    ph = "pra" if f"pra_{kind}" not in found else "nda"
                if f"{ph}_{kind}" not in found:
                    found[f"{ph}_{kind}"] = c
        if len(found) > len(info):
            info, header_row = found, r

    if "project" not in info:
        raise ValueError(
            f"양식에서 Project 컬럼을 찾지 못했습니다 (월 헤더 {month_row}행). "
            "config.json의 template_aliases를 확인하세요."
        )
    if not spans:
        notes.append("PRA/NDA 그룹 제목을 못 찾아 Sub/App을 왼쪽부터 순서대로 배정했습니다.")

    first_data = max(header_row, month_row) + 1
    last_data = first_data - 1
    for r in range(first_data, (ws.max_row or first_data) + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in info.values()):
            last_data = r

    return TemplateLayout(ws.title, header_row, first_data, last_data,
                          month_cols, info, notes)


def find_template_sheet(wb, config: dict,
                        preferred: str | None = None) -> tuple[Worksheet, TemplateLayout]:
    names = [preferred] if preferred and preferred in wb.sheetnames else wb.sheetnames
    errs = []
    for name in names:
        try:
            ws = wb[name]
            return ws, detect_template_layout(ws, config)
        except ValueError as e:
            errs.append(f"  · {name}: {e}")
    raise ValueError("간트 양식 시트를 찾지 못했습니다.\n" + "\n".join(errs))


# ───────────────────────────────────────────────────────────────────────
# 색상 — 만들어내지 않고, 양식에서 배운다
# ───────────────────────────────────────────────────────────────────────
def learn_palette(ws: Worksheet, layout: TemplateLayout) -> dict[str, dict[str, PatternFill]]:
    """
    양식에 이미 그려져 있는 바에서 색을 **그대로 배운다**.

    색은 함부로 바꾸면 안 된다 — 사용자 팀의 규칙이 담겨 있다. 실측 결과
    색조(hue)는 Category가 정하고, 진하기(tint)가 준비/검토를 가른다.
    그러니 Category별로 '제출 이전에 쓰인 색'과 '제출~승인에 쓰인 색'을
    세어 가장 많이 쓰인 것을 그 Category의 색으로 삼는다.

    반환: {category_lower: {"prep": PatternFill, "review": PatternFill}}
    """
    info = layout.info_cols
    if "category" not in info:
        return {}

    def fill_key(cell):
        f = cell.fill
        if f is None or f.fill_type is None:
            return None
        c = f.fgColor
        if c is None:
            return None
        if c.type == "rgb" and isinstance(c.rgb, str) and c.rgb != "00000000":
            return ("rgb", c.rgb, 0.0)
        if c.type == "theme":
            return ("theme", c.theme, round(c.tint or 0.0, 3))
        if c.type == "indexed":
            return ("indexed", c.indexed, 0.0)
        return None

    from collections import Counter, defaultdict
    tally: dict[str, dict[str, Counter]] = defaultdict(
        lambda: {"prep": Counter(), "review": Counter()})

    for r in range(layout.first_data_row, layout.last_data_row + 1):
        cat = _low(ws.cell(row=r, column=info["category"]).value)
        if not cat:
            continue
        subs, apps = [], []
        for ph in PHASES:
            for kind, bucket in (("sub", subs), ("app", apps)):
                c = info.get(f"{ph}_{kind}")
                if c:
                    v = ws.cell(row=r, column=c).value
                    if hasattr(v, "year"):
                        bucket.append(v.date() if hasattr(v, "date") else v)
        if not subs:
            continue
        sub, app = min(subs), (max(apps) if apps else None)
        for c, m in layout.month_cols.items():
            key = fill_key(ws.cell(row=r, column=c))
            if not key:
                continue
            if app and sub <= m <= app:
                tally[cat]["review"][key] += 1
            elif m < sub:
                tally[cat]["prep"][key] += 1

    def to_fill(key):
        from openpyxl.styles.colors import Color
        kind, val, tint = key
        if kind == "rgb":
            return PatternFill("solid", fgColor=Color(rgb=val))
        if kind == "theme":
            return PatternFill("solid", fgColor=Color(theme=val, tint=tint))
        return PatternFill("solid", fgColor=Color(indexed=val))

    palette: dict[str, dict[str, PatternFill]] = {}
    for cat, phases in tally.items():
        entry = {}
        for phase in ("prep", "review"):
            if phases[phase]:
                entry[phase] = to_fill(phases[phase].most_common(1)[0][0])
        if entry:
            palette[cat] = entry
    return palette


def _spec_to_fill(spec: dict) -> PatternFill | None:
    """config의 색 명세({theme,tint} 또는 {rgb})를 PatternFill로."""
    from openpyxl.styles.colors import Color
    if not spec:
        return None
    if "theme" in spec:
        return PatternFill("solid", fgColor=Color(theme=int(spec["theme"]),
                                                  tint=float(spec.get("tint", 0.0))))
    if "rgb" in spec:
        return PatternFill("solid", fgColor=Color(rgb=spec["rgb"]))
    if "indexed" in spec:
        return PatternFill("solid", fgColor=Color(indexed=int(spec["indexed"])))
    return None


def config_palette(config: dict) -> dict[str, dict[str, PatternFill]]:
    """
    config.json에 구워둔 Category별 실제 색.

    내장 양식은 데이터가 비워져 있어 배울 바가 없다. 그래서 실물 파일에서
    추출해 둔 이 값을 쓴다 (tools/make_blank_template.py로 뽑은 것).
    """
    out: dict[str, dict[str, PatternFill]] = {}
    for cat, phases in config.get("category_palette", {}).items():
        entry = {}
        for phase in ("prep", "review"):
            fill = _spec_to_fill(phases.get(phase, {}))
            if fill:
                entry[phase] = fill
        if entry:
            out[cat.lower()] = entry
    return out


# 엑셀 테마 색 인덱스 → theme1.xml의 요소 이름.
# openpyxl의 fgColor.theme 값이 이 순서를 따른다.
_THEME_SLOTS = ["lt1", "dk1", "lt2", "dk2",
                "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
                "hlink", "folHlink"]

# Office 기본 테마 (통합문서에서 못 읽었을 때의 예비값)
_DEFAULT_THEME_RGB = {
    "lt1": "FFFFFF", "dk1": "000000", "lt2": "E7E6E6", "dk2": "44546A",
    "accent1": "4472C4", "accent2": "ED7D31", "accent3": "A5A5A5",
    "accent4": "FFC000", "accent5": "5B9BD5", "accent6": "70AD47",
    "hlink": "0563C1", "folHlink": "954F72",
}


def read_theme_colors(wb) -> dict[str, str]:
    """
    통합문서의 theme1.xml에서 테마 색 12개를 읽는다.

    바 색이 테마 색으로 저장돼 있어서(예: theme9+tint0.6) 미리보기에 실제
    색을 보여주려면 이 표가 필요하다. 못 읽으면 Office 기본 테마를 쓴다.
    """
    try:
        theme_xml = wb.loaded_theme
        if isinstance(theme_xml, bytes):
            theme_xml = theme_xml.decode("utf-8", "ignore")
        if not theme_xml:
            return dict(_DEFAULT_THEME_RGB)
    except AttributeError:
        return dict(_DEFAULT_THEME_RGB)

    out = dict(_DEFAULT_THEME_RGB)
    # <a:accent3><a:srgbClr val="A5A5A5"/></a:accent3> 형태를 훑는다
    for slot in _THEME_SLOTS:
        m = re.search(
            rf"<a:{slot}>\s*<a:(srgbClr|sysClr)[^>]*?"
            rf"(?:val|lastClr)=\"([0-9A-Fa-f]{{6}})\"", theme_xml)
        if m:
            out[slot] = m.group(2).upper()
    return out


def apply_tint(rgb: str, tint: float) -> str:
    """
    엑셀의 tint를 적용해 최종 RGB를 얻는다.

    엑셀은 HSL의 명도(luminance)에 tint를 건다.
      tint > 0 : L' = L*(1-tint) + tint   (밝게)
      tint < 0 : L' = L*(1+tint)          (어둡게)
    이 규칙이 있어야 '연한 색=준비 / 진한 색=검토'가 화면에서도 재현된다.
    """
    import colorsys

    rgb = rgb[-6:]          # 'FFD3D7DC' 같은 ARGB에서 알파를 떼어낸다
    if not tint:
        return rgb.upper()
    r, g, b = (int(rgb[i:i + 2], 16) / 255 for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = l * (1 - tint) + tint if tint > 0 else l * (1 + tint)
    l = max(0.0, min(1.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return "{:02X}{:02X}{:02X}".format(round(r * 255), round(g * 255), round(b * 255))


def fill_to_css(fill: PatternFill, theme_rgb: dict[str, str]) -> str | None:
    """PatternFill을 '#RRGGBB' 문자열로. 미리보기가 실제 색을 그대로 보여주기 위한 것."""
    if fill is None or fill.fill_type is None:
        return None
    c = fill.fgColor
    if c is None:
        return None
    if c.type == "rgb" and isinstance(c.rgb, str) and c.rgb != "00000000":
        return "#" + apply_tint(c.rgb, c.tint or 0.0)
    if c.type == "theme":
        try:
            slot = _THEME_SLOTS[int(c.theme)]
        except (IndexError, ValueError, TypeError):
            return None
        return "#" + apply_tint(theme_rgb.get(slot, "999999"), c.tint or 0.0)
    return None


def palette_to_css(palette: dict, theme_rgb: dict[str, str]) -> dict[str, dict[str, str]]:
    """Category별 팔레트를 화면용 색 표로. {category: {prep:'#..', review:'#..'}}"""
    out: dict[str, dict[str, str]] = {}
    for cat, phases in palette.items():
        entry = {}
        for phase, fill in phases.items():
            css = fill_to_css(fill, theme_rgb)
            if css:
                entry[phase] = css
        if entry:
            out[cat] = entry
    return out


def _config_fills(status: str, category: str, config: dict) -> tuple[PatternFill, PatternFill]:
    """팔레트에도 없는 Category일 때만 쓰는 최후의 예비 색."""
    for k, v in config.get("category_colors", {}).items():
        if k.lower() == category.lower():
            return PatternFill("solid", fgColor=v[0]), PatternFill("solid", fgColor=v[1])
    for k, v in config.get("status_colors", {}).items():
        if k.lower() == status.lower():
            return PatternFill("solid", fgColor=v[0]), PatternFill("solid", fgColor=v[1])
    d = config.get("default_status_color", ["EEEEEE", "999999"])
    return PatternFill("solid", fgColor=d[0]), PatternFill("solid", fgColor=d[1])


def draw_prep_period(config: dict) -> bool:
    """준비기간(연한 색) 바를 그릴 것인가. 기본은 그리지 않는다."""
    return bool(config.get("draw_prep_period", False))


def resolve_fills(rec: Record, config: dict,
                  palette: dict) -> tuple[PatternFill, PatternFill]:
    """
    이 행을 칠할 색 한 쌍(준비, 검토)을 정한다.

    찾는 순서가 중요하다. 간트에 적히는 Category는 **RA plan 원본 그대로**인데,
    내장 양식의 팔레트는 접힌 이름(color_group, 예: Others)으로 배워져 있다.
    그래서 원본 Category로 먼저 찾고, 없으면 색 그룹으로 찾는다.

      · 내장 빈 양식     → 원본 이름 없음 → 색 그룹으로 적중 (기존 색 유지)
      · 이 앱이 만든 간트를 다시 양식으로 올림
                        → 원본 이름으로 적중 (사용자가 칠한 색을 그대로 배운다)

    이 함수 하나로 엑셀 채색과 화면 미리보기가 같은 답을 쓰게 만든다.
    두 군데서 따로 계산하면 언젠가 반드시 어긋난다.
    """
    learned = (palette.get(_low(rec.category))
               or palette.get(_low(rec.color_group))
               or {})

    # 예비 색도 같은 순서로 찾는다. _config_fills는 못 찾아도 상태색·기본색을
    # 돌려주므로 "못 찾았다"를 반환값으로는 알 수 없다. 그래서 category_colors에
    # 실제로 키가 있는지를 먼저 보고 고른다.
    known = {k.lower() for k in config.get("category_colors", {})}
    fb_cat = (rec.category if _low(rec.category) in known
              else rec.color_group)
    fb_prep, fb_review = _config_fills(rec.status, fb_cat, config)
    return learned.get("prep", fb_prep), learned.get("review", fb_review)


def _paint_row(ws: Worksheet, row: int, rec: Record, month_cols: dict[int, date],
               config: dict, palette: dict) -> None:
    """
    한 행의 월 셀을 칠한다.

    PRA와 NDA 두 구간을 각각 그린다. 겹치면 검토(진한색)가 준비(연한색)를
    덮는다 — 진행된 사실이 계획보다 우선이기 때문.

    준비기간은 기본적으로 그리지 않는다(`draw_prep_period`). 그 길이는 실측이
    아니라 규제 관행에서 뽑은 추정치라, 그려 놓으면 사람이 확정된 일정으로
    읽는다. 확실하지 않은 것을 확실한 것처럼 보이게 하느니 비워 둔다.
    """
    prep_fill, review_fill = resolve_fills(rec, config, palette)
    with_prep = draw_prep_period(config)

    spans_prep, spans_review = [], []
    for ph in PHASES:
        sub = getattr(rec, f"{ph}_sub")
        app = rec.app_or_est(ph)               # 승인일 없으면 표준 검토기간으로 추정
        prep = getattr(rec, f"{ph}_prep")
        if with_prep and prep and sub:
            spans_prep.append((prep, sub))
        if sub and app:
            spans_review.append((sub, app))
        elif sub:
            spans_review.append((sub, sub))     # 추정조차 못 하면 제출월만 마커

    for c, m in month_cols.items():
        m_end = add_months(m, 1)
        cell = ws.cell(row=row, column=c)
        fill = None
        if any(s < m_end and e > m for s, e in spans_prep):
            fill = prep_fill
        if any(s < m_end and (e > m or month_floor(s) == m) for s, e in spans_review):
            fill = review_fill
        cell.fill = fill if fill else PatternFill(fill_type=None)


# ───────────────────────────────────────────────────────────────────────
# 정렬 · 변경 감지
# ───────────────────────────────────────────────────────────────────────
def record_key(rec: Record) -> str:
    """행의 신원. 행 번호는 팀원이 중간에 끼워 넣으면 밀리므로 제품+과제명으로 본다."""
    return f"{_low(rec.product)}|{_low(rec.project)}"


def sort_records(records: list[Record], config: dict) -> list[Record]:
    order = config.get("sort_by", ["responsible", "category", "nda_sub"])
    far = date(9999, 12, 31)

    def key(r: Record):
        out = []
        for f_ in order:
            v = getattr(r, f_, "") if hasattr(r, f_) else ""
            if isinstance(v, date):
                out.append((0, "", v))
            elif v in (None, ""):
                out.append((1, "", far))
            else:
                out.append((0, str(v).lower(), far))
        return out

    return sorted(records, key=key)


DATE_FIELDS = ("pra_sub", "pra_app", "nda_sub", "nda_app")


def diff_records(previous: list[dict], current: list[Record]) -> dict:
    """
    이전 간트 대비 무엇이 달라졌는지. 이 도구의 존재 이유.

    매니저가 알고 싶은 것은 '간트 파일'이 아니라 '지난번 이후 무엇이 바뀌었나'다.
    """
    prev = {f"{_low(p.get('product'))}|{_low(p.get('project'))}": p for p in previous}
    seen, added, moved, status_changed = set(), [], [], []

    for rec in current:
        k = record_key(rec)
        seen.add(k)
        old = prev.get(k)
        if old is None:
            added.append(rec.to_dict())
            continue
        changes = {}
        for f_ in DATE_FIELDS:
            o = old.get(f_)
            n = getattr(rec, f_)
            n = n.isoformat() if n else None
            if o != n:
                changes[f_] = {"from": o, "to": n}
        if changes:
            moved.append({**rec.to_dict(), "date_changes": changes})
        if _norm(old.get("status")) != _norm(rec.status):
            status_changed.append({**rec.to_dict(),
                                   "from_status": old.get("status"),
                                   "to_status": rec.status})

    removed = [p for k, p in prev.items() if k not in seen]
    return {
        "added": added, "moved": moved,
        "status_changed": status_changed, "removed": removed,
        "has_previous": bool(previous),
        "summary": {"added": len(added), "moved": len(moved),
                    "status_changed": len(status_changed), "removed": len(removed)},
    }


def read_gantt_rows(path: str | Path, config: dict,
                    sheet_name: str | None = None) -> list[dict]:
    """기존 간트에 실려 있는 내용(변경 감지의 '이전' 쪽). 비어 있으면 전부 신규가 된다."""
    try:
        wb = load_workbook(path, data_only=True)
        ws, layout = find_template_sheet(wb, config, sheet_name)
    except (ValueError, KeyError, OSError):
        return []

    info = layout.info_cols
    rows = []
    for r in range(layout.first_data_row, layout.last_data_row + 1):
        def g(key):
            c = info.get(key)
            return ws.cell(row=r, column=c).value if c else None

        project = _norm(g("project"))
        if not project:
            continue
        row = {"responsible": _norm(g("responsible")), "category": _norm(g("category")),
               "product": _norm(g("product")), "project": project,
               "status": _norm(g("status"))}
        for f_ in DATE_FIELDS:
            v = g(f_)
            v = v.date() if hasattr(v, "date") else (v if hasattr(v, "year") else None)
            row[f_] = v.isoformat() if v else None
        rows.append(row)
    return rows


# ───────────────────────────────────────────────────────────────────────
# 쓰기
# ───────────────────────────────────────────────────────────────────────
def write_into_template(template_path, records: list[Record], output_path,
                        config: dict, sheet_name: str | None = None) -> dict:
    """
    사용자 양식에 내용만 채운다. 서식·열너비·병합·색 규칙은 그대로 둔다.

    원본 파일은 건드리지 않는다(복사본에 쓴다). 결과 시트는 날짜 이름으로
    바꿔 기존의 '개정 때마다 탭 복제' 관행을 그대로 이어받는다.
    """
    shutil.copyfile(template_path, output_path)
    wb = load_workbook(output_path)
    ws, layout = find_template_sheet(wb, config, sheet_name)

    # 색은 지어내지 않는다. config에 구워둔 실측 팔레트를 깔고, 양식에 바가
    # 남아 있으면 그쪽에서 배운 색으로 덮는다 (양식이 언제나 최종 권위).
    palette = config_palette(config)
    palette.update(learn_palette(ws, layout))

    stamp = datetime.now().strftime(config.get("snapshot_tab_naming", "%Y%m%d"))
    ws.title = stamp if stamp not in wb.sheetnames else f"{stamp}_{datetime.now():%H%M}"

    info = layout.info_cols
    all_cols = list(info.values()) + list(layout.month_cols)

    # 서식 견본을 확보한 뒤 기존 데이터를 비운다
    style: dict[int, tuple] = {}
    if layout.last_data_row >= layout.first_data_row:
        for c in all_cols:
            s = ws.cell(row=layout.first_data_row, column=c)
            style[c] = (copy(s.font), copy(s.border), copy(s.alignment), s.number_format)
    for r in range(layout.first_data_row, layout.last_data_row + 1):
        for c in all_cols:
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

    getters = {
        # Category는 RA plan에 적힌 것을 그대로 쓴다. 접힌 이름(category)은
        # 색을 고를 때만 쓰고 셀에는 쓰지 않는다 — 팀원이 자기가 적은 값을
        # 간트에서 그대로 찾을 수 있어야 하기 때문이다.
        "responsible": lambda x: x.responsible,
        "category": lambda x: x.category,
        "product": lambda x: x.product, "project": lambda x: x.project,
        "status": lambda x: x.status,
        "pra_sub": lambda x: x.pra_sub, "pra_app": lambda x: x.pra_app,
        "nda_sub": lambda x: x.nda_sub, "nda_app": lambda x: x.nda_app,
    }

    # 원본 양식에는 필터가 걸린 채로 저장돼 숨겨진 행이 남아 있다(실물 180행).
    # 그대로 두면 새로 쓴 행이 엑셀에서 보이지 않으므로 데이터 구간을 전부 편다.
    last_written_row = layout.first_data_row + max(len(records), 0) - 1
    for r in range(layout.first_data_row,
                   max(last_written_row, layout.last_data_row) + 1):
        if ws.row_dimensions[r].hidden:
            ws.row_dimensions[r].hidden = False

    # 행만 펴는 걸로는 부족하다. 필터 '조건'이 양식에 그대로 남아 있으면
    # (실물: A열에 특정 값 하나만 보이도록 걸려 있었다) 파일은 멀쩡히 열리지만
    # 사용자가 필터를 한 번 건드리는 순간 엑셀이 조건을 다시 적용해 행이 도로
    # 사라진다. 조건은 지우고, 필터 범위는 이번에 쓴 데이터에 다시 맞춘다.
    # 필터 자체는 없애지 않는다 — 매니저가 실제로 쓰는 기능이라서다.
    if ws.auto_filter is not None:
        ws.auto_filter.filterColumn = []
        ws.auto_filter.sortState = None
        if ws.auto_filter.ref and info:
            first_col, last_col = min(info.values()), max(info.values())
            ws.auto_filter.ref = (
                f"{get_column_letter(first_col)}{layout.header_row}"
                f":{get_column_letter(last_col)}{max(last_written_row, layout.header_row)}"
            )

    for i, rec in enumerate(records):
        row = layout.first_data_row + i
        for key, col in info.items():
            cell = ws.cell(row=row, column=col)
            cell.value = getters.get(key, lambda _: None)(rec)
            if col in style:
                f, b, a, nf = style[col]
                cell.font, cell.border, cell.alignment = copy(f), copy(b), copy(a)
                cell.number_format = nf
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"
        for col in layout.month_cols:
            if col in style:
                ws.cell(row=row, column=col).border = copy(style[col][1])
        _paint_row(ws, row, rec, layout.month_cols, config, palette)

    wb.save(output_path)
    months = sorted(layout.month_cols.values())

    # 미리보기가 엑셀과 똑같은 색을 보여줄 수 있도록 팔레트를 화면용 색으로 변환.
    # 키는 화면에 실제로 찍히는 Category(원본 그대로)로 맞춘다. 채색과 같은
    # resolve_fills를 써서 뽑으므로 미리보기 색과 엑셀 색이 어긋날 수 없다.
    # 이 간트에 실제로 등장하는 Category만 담는다. palette_to_css로 팔레트를
    # 통째로 넘기면 간트 양식이 쓰던 접힌 이름(Others, NDA (variant), Device)이
    # 범례에 섞여 나온다 — RA plan에는 없는 이름이라 보는 사람이 혼란스럽다.
    theme_rgb = read_theme_colors(wb)
    css_palette: dict[str, dict[str, str]] = {}
    for rec in records:
        cat = _low(rec.category)
        if not cat or cat in css_palette:
            continue
        p, rv = resolve_fills(rec, config, palette)
        css_palette[cat] = {"prep": fill_to_css(p, theme_rgb) or "#DDDDDD",
                            "review": fill_to_css(rv, theme_rgb) or "#999999"}

    return {
        "mode": "template",
        "sheet": ws.title,
        "rows_written": len(records),
        "month_span": [months[0].isoformat(), months[-1].isoformat()],
        "palette_learned": sorted(palette.keys()),
        "palette_css": css_palette,
        "template_notes": layout.notes,
        "info_cols": {k: get_column_letter(v) for k, v in info.items()},
    }


def generate(ra_plan_path, output_path, template_path=None, config: dict | None = None,
             sheet_name: str | None = None, previous_rows: list[dict] | None = None) -> dict:
    """
    RA plan → 간트 엑셀. 이 도구의 본체.

    전량 재생성이다. 일부만 고쳐 넣지 않고 RA plan을 통째로 다시 읽어 간트를
    새로 그린다. 그래야 사람이 동기화를 신경 쓸 여지가 아예 없어진다.
    """
    config = config or load_config()
    if not template_path:
        raise ValueError(
            "간트 양식(템플릿) 파일이 필요합니다. 사용자 양식에만 맞춰 그리도록 "
            "설계돼 있어 양식 없이는 생성하지 않습니다."
        )

    records, cols, source_sheet = read_ra_plan(ra_plan_path, config, sheet_name)
    dated = [r for r in records if r.has_dates]

    # 이미 끝난 옛날 건은 간트에 올리지 않는다. 타임라인 창(2025~) 밖이라
    # 어차피 바가 안 그려지고, 빈 행만 수백 개 쌓여 읽기가 어려워진다.
    # 판정은 '마지막 시점'으로 하되 추정 승인일까지 포함한다 — 2024년에
    # 제출하고 아직 심사 중인 건은 살아 있는 건이므로 남겨야 한다.
    cutoff_raw = config.get("min_activity_date")
    cutoff = date.fromisoformat(cutoff_raw) if cutoff_raw else None
    if cutoff:
        kept = [r for r in dated
                if r.last_activity is None or r.last_activity >= cutoff]
    else:
        kept = dated
    records_filtered = len(dated) - len(kept)
    usable = sort_records(kept, config)

    if previous_rows is None:
        previous_rows = read_gantt_rows(template_path, config, sheet_name)
    changes = diff_records(previous_rows or [], usable)

    result = write_into_template(template_path, usable, output_path, config, sheet_name)
    result.update({
        "source_sheet": source_sheet,
        "header_row": cols.header_row,
        "records_total": len(records),
        "records_written": len(usable),
        "records_skipped": len(records) - len(dated),
        "records_filtered": records_filtered,
        "min_activity_date": cutoff_raw,
        "detect_notes": cols.notes,
        "records": [r.to_dict() for r in usable],
        "changes": changes,
    })
    return result


# ───────────────────────────────────────────────────────────────────────
# calibrate — 추측을 실측으로 갈음한다
# ───────────────────────────────────────────────────────────────────────
def calibrate(ra_plan_path, existing_gantt_path, config: dict | None = None) -> dict:
    """
    실물 RA plan과 실물 간트를 나란히 놓고 규칙을 역산한다.

      · 간트의 각 날짜 컬럼이 RA plan 어느 컬럼에서 왔는지 (값 집합 대조)
      · Category 원본 → 간트 표시값 실측 대응
      · Category별 색상 (양식에서 배운 팔레트)

    쓰는 법:  python gantt_engine.py calibrate <RA plan.xlsx> <실제간트.xlsx>
    결과를 보고 config.json을 고치면 된다. 코드는 고칠 필요 없다.
    """
    from collections import Counter, defaultdict

    config = config or load_config()
    tokens = config.get("unresolved_tokens", ["TBD"])

    records, cols, _ = read_ra_plan(ra_plan_path, config)
    wb = load_workbook(ra_plan_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for name in wb.sheetnames:
        try:
            detect_source_columns(wb[name], config)
            ws = wb[name]
            break
        except ValueError:
            continue
    cols = detect_source_columns(ws, config)

    plan_sets: dict[str, set] = {}
    for ph in PHASES:
        for kind in ("sub", "app"):
            for which in ("planned", "actual"):
                attr = f"{ph}_{which}_{kind}"
                c = getattr(cols, attr)
                s = set()
                if c:
                    for r in range(cols.header_row + 1, (ws.max_row or 0) + 1):
                        dv = parse_cell_date(ws.cell(row=r, column=c).value, tokens)
                        if dv.usable and not dv.partial:
                            s.add(dv.value)
                plan_sets[attr] = s

    gwb = load_workbook(existing_gantt_path, data_only=True)
    gws, layout = find_template_sheet(gwb, config)
    info = layout.info_cols

    coverage: dict[str, dict] = {}
    for ph in PHASES:
        for kind in ("sub", "app"):
            gcol = info.get(f"{ph}_{kind}")
            if not gcol:
                continue
            vals = []
            for r in range(layout.first_data_row, layout.last_data_row + 1):
                v = gws.cell(row=r, column=gcol).value
                v = v.date() if hasattr(v, "date") else (v if hasattr(v, "year") else None)
                if v:
                    vals.append(v)
            if not vals:
                coverage[f"{ph}_{kind}"] = {"values": 0, "note": "이 쌍은 비어 있음"}
                continue
            scores = {
                attr: round(sum(1 for v in vals if v in s) / len(vals) * 100, 1)
                for attr, s in plan_sets.items() if attr.endswith(kind)
            }
            coverage[f"{ph}_{kind}"] = {
                "values": len(vals),
                "match_pct": dict(sorted(scores.items(), key=lambda x: -x[1])),
            }

    cat_obs: dict[str, Counter] = defaultdict(Counter)
    if "category" in info:
        for r in range(layout.first_data_row, layout.last_data_row + 1):
            cat_obs["(간트 표시값)"][_norm(gws.cell(row=r, column=info["category"]).value)] += 1

    palette = learn_palette(gws, layout)
    return {
        "ra_plan_records": len(records),
        "gantt_rows": layout.last_data_row - layout.first_data_row + 1,
        "column_coverage": coverage,
        "gantt_category_distribution": {k: dict(v) for k, v in cat_obs.items()},
        "palette_learned_categories": sorted(palette.keys()),
        "gantt_month_span": [min(layout.month_cols.values()).isoformat(),
                             max(layout.month_cols.values()).isoformat()],
        "gantt_info_cols": {k: get_column_letter(v) for k, v in info.items()},
        "hint": (
            "column_coverage의 match_pct에서 1위가 그 컬럼의 출처입니다. "
            "actual과 planned가 비슷하면 '실제값 우선, 없으면 계획값' 규칙이 맞는 것입니다. "
            "Project 컬럼이 비어 있는 익명화 사본이면 행 단위 대조는 불가하고 "
            "값 집합 대조만 유효합니다."
        ),
    }


# ───────────────────────────────────────────────────────────────────────
# CLI
# ───────────────────────────────────────────────────────────────────────
def _main() -> int:
    import argparse

    p = argparse.ArgumentParser(prog="gantt_engine",
                                description="RA plan 엑셀 → 간트 엑셀 (LLM 없이 순수 코드)")
    s = p.add_subparsers(dest="cmd", required=True)

    g = s.add_parser("generate", help="간트 엑셀 생성")
    g.add_argument("ra_plan")
    g.add_argument("template", help="사용자 간트 양식 파일 (필수)")
    g.add_argument("-o", "--output", default="gantt_output.xlsx")
    g.add_argument("-s", "--sheet", default=None)
    g.add_argument("-c", "--config", default=None)

    d = s.add_parser("inspect", help="판독 결과만 출력")
    d.add_argument("ra_plan")
    d.add_argument("-c", "--config", default=None)

    c = s.add_parser("calibrate", help="실제 간트에서 규칙 역산")
    c.add_argument("ra_plan")
    c.add_argument("existing_gantt")
    c.add_argument("-c", "--config", default=None)

    a = p.parse_args()
    cfg = load_config(a.config)

    if a.cmd == "generate":
        res = generate(a.ra_plan, a.output, a.template, cfg, a.sheet)
        recs = res.pop("records", [])
        print(json.dumps(res, ensure_ascii=False, indent=2))
        print(f"\n{len(recs)}건 기록 → {a.output}")

    elif a.cmd == "inspect":
        records, cols, sheet = read_ra_plan(a.ra_plan, cfg)
        print(f"시트 {sheet} / 헤더 {cols.header_row}행 / 레코드 {len(records)}건 "
              f"(날짜 있음 {sum(1 for r in records if r.has_dates)}건)")
        for n in cols.notes:
            print("  ·", n)
        print()
        for r in [x for x in records if x.has_dates][:30]:
            print(f"[{r.source_row:>4}] {r.responsible:<6} {r.category:<15} {r.product:<6} "
                  f"PRA {str(r.pra_sub or '-'):<11}→{str(r.pra_app or '-'):<11} "
                  f"NDA {str(r.nda_sub or '-'):<11}→{str(r.nda_app or '-'):<11} {r.status}")

    elif a.cmd == "calibrate":
        print(json.dumps(calibrate(a.ra_plan, a.existing_gantt, cfg),
                         ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
