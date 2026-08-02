# -*- coding: utf-8 -*-
"""
실물 간트 파일에서 **빈 양식**을 뽑아낸다.

왜 필요한가
─────────────────────────────────────────────────────────────────────────
사용자는 매번 양식을 올리고 싶어 하지 않는다. 양식은 이미 정해져 있으니
프로젝트에 내장돼 있어야 한다. 그러나 실물 간트 파일에는 제품명·과제명·
일정이 들어 있어(Internal Use) 저장소에 넣을 수 없다.

그래서 이 스크립트가 하는 일:
  · 서식은 전부 남긴다 — 색, 글꼴, 열너비, 행높이, 병합, 틀고정, Note 문구,
    연도/월 2단 헤더, 조건부서식, 인쇄설정
  · 데이터 행의 **값만** 지운다 (셀 스타일은 그대로 둔다)
  · 동시에 바 색상 팔레트를 Category별로 학습해 JSON으로 뽑는다
    → 빈 양식에는 배울 바가 없으므로, 이 값을 config.json에 구워 넣어야
      색이 원래대로 재현된다

    python tools/make_blank_template.py <실물간트.xlsx> [-o template/RA_gantt_template.xlsx]
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import warnings
warnings.filterwarnings("ignore")

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import gantt_engine as engine


def extract_palette_spec(ws, layout) -> dict:
    """learn_palette가 배운 PatternFill을 config.json에 넣을 수 있는 형태로 바꾼다."""
    palette = engine.learn_palette(ws, layout)
    out: dict[str, dict] = {}
    for cat, phases in palette.items():
        entry = {}
        for phase, fill in phases.items():
            c = fill.fgColor
            if c.type == "theme":
                entry[phase] = {"theme": c.theme, "tint": round(c.tint or 0.0, 3)}
            elif c.type == "rgb" and isinstance(c.rgb, str):
                entry[phase] = {"rgb": c.rgb}
            elif c.type == "indexed":
                entry[phase] = {"indexed": c.indexed}
        if entry:
            out[cat] = entry
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("source", help="실물 간트 엑셀")
    ap.add_argument("-o", "--output", default="template/RA_gantt_template.xlsx")
    ap.add_argument("-p", "--palette-out", default="template/palette.json")
    ap.add_argument("-s", "--sheet", default=None)
    args = ap.parse_args()

    config = engine.load_config()
    src = Path(args.source)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)

    # 1) 팔레트 학습은 원본에서 (값이 살아 있어야 배울 수 있다)
    wb_read = load_workbook(src, data_only=True)
    ws_read, layout = engine.find_template_sheet(wb_read, config, args.sheet)
    palette_spec = extract_palette_spec(ws_read, layout)
    print(f"학습한 Category 색상 {len(palette_spec)}종: {', '.join(sorted(palette_spec))}")

    # 2) 서식 유지본을 복사한 뒤 값만 비운다
    shutil.copyfile(src, out)
    wb = load_workbook(out)
    ws, layout2 = engine.find_template_sheet(wb, config, args.sheet)

    # 데이터 행은 '아는 컬럼만' 지우면 안 된다 — 반드시 전부 지운다.
    #
    # 처음엔 info_cols(A~I)만 지웠다. 그랬더니 타임라인 오른쪽 끝 Note 열(BF)에
    # 담당자가 손으로 적어 둔 내부 코멘트 25건이 그대로 살아남아 저장소에
    # 올라갔다("202606 재제출 예정", "NNPK가 안할 가능성 있음" 같은 것들).
    # 어느 열에 사람이 무엇을 적어 뒀는지는 미리 알 수 없다. 그러니 아는 것만
    # 지우는 화이트리스트가 아니라, 데이터 구간은 통째로 비우고 서식만 남긴다.
    cleared = 0
    last_col = ws.max_column or max(layout2.info_cols.values())
    for r in range(layout2.first_data_row, layout2.last_data_row + 1):
        for c in range(1, last_col + 1):
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None     # 스타일은 건드리지 않는다
                cleared += 1
        for c in layout2.month_cols:                     # 바 색만 제거, 테두리는 유지
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)

    # 데이터가 들어 있던 다른 시트도 값을 비운다 (Sheet 2 등)
    for name in wb.sheetnames:
        if name == ws.title:
            continue
        other = wb[name]
        for row in other.iter_rows():
            for cell in row:
                cell.value = None

    wb.save(out)
    print(f"빈 양식 저장: {out}  (값 {cleared}개 제거, 서식 유지)")

    Path(args.palette_out).write_text(
        json.dumps(palette_spec, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"팔레트 저장: {args.palette_out}")
    print("\n→ 이 팔레트를 config.json의 'category_palette'에 넣으면 빈 양식에서도 "
          "원래 색이 재현됩니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
